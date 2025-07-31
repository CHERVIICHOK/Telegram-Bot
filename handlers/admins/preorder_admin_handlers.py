from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, BufferedInputFile, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.fsm.context import FSMContext
from datetime import datetime
import os
import logging
import openpyxl
from openpyxl import Workbook
from io import BytesIO

from database.preorder_db import preorder_db
from keyboards.admins.preorder_admin_keyboards import (
    get_preorder_admin_menu_keyboard,
    get_preorder_products_list_keyboard,
    get_product_admin_keyboard,
    get_confirm_delete_keyboard,
    get_add_product_cancel_keyboard,
    get_confirm_add_keyboard,
    get_skip_step_keyboard,
    get_category_selection_keyboard,
    get_product_name_selection_keyboard,
    get_edit_field_keyboard,
    get_stats_keyboard,
    get_bulk_upload_keyboard,
    get_bulk_upload_confirm_keyboard
)
from states.preorder_states import AddPreorderProduct, BulkUploadProducts
from filters.admin_filter import AdminFilter

logger = logging.getLogger(__name__)
router = Router()

router.message.filter(AdminFilter())
router.callback_query.filter(AdminFilter())


@router.callback_query(F.data == "manage_preorders")
async def show_preorder_admin_menu(callback: CallbackQuery):
    """Показать меню управления предзаказами"""
    await callback.message.edit_text(
        "🛍️ <b>Управление предзаказами</b>\n\n"
        "Выберите действие:",
        reply_markup=get_preorder_admin_menu_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "admin:main")
async def back_to_admin_menu(callback: CallbackQuery):
    """Вернуться в главное админ-меню"""
    # Импортируем здесь, чтобы избежать циклических импортов
    from keyboards.admins.menu_keyboard import get_admin_menu_keyboard

    await callback.message.edit_text(
        "⚙️ <b>Панель администратора</b>\n\n"
        "Выберите нужную функцию:",
        reply_markup=get_admin_menu_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:menu")
async def back_to_preorder_menu(callback: CallbackQuery):
    """Вернуться в меню предзаказов"""
    await show_preorder_admin_menu(callback)


@router.callback_query(F.data == "preorder_admin:add")
async def start_add_product(callback: CallbackQuery, state: FSMContext):
    """Начать добавление товара"""
    await state.set_state(AddPreorderProduct.category)

    # Получаем существующие категории
    categories = preorder_db.get_all_categories()

    if categories:
        await callback.message.edit_text(
            "Выберите существующую категорию или введите новую:",
            reply_markup=get_category_selection_keyboard(categories)
        )
    else:
        await callback.message.edit_text(
            "Введите категорию товара:",
            reply_markup=get_add_product_cancel_keyboard()
        )

    await callback.answer()


@router.callback_query(F.data.startswith("preorder_admin:select_category:"), AddPreorderProduct.category)
async def select_existing_category(callback: CallbackQuery, state: FSMContext):
    """Выбор существующей категории"""
    category = callback.data.split(":", 3)[2]
    await state.update_data(category=category)
    await state.set_state(AddPreorderProduct.product_name)

    # Получаем существующие названия товаров
    product_names = preorder_db.get_all_product_names()

    if product_names:
        await callback.message.edit_text(
            f"Категория: {category}\n\nВыберите существующее название или введите новое:",
            reply_markup=get_product_name_selection_keyboard(product_names)
        )
    else:
        await callback.message.edit_text(
            f"Категория: {category}\n\nВведите название товара:",
            reply_markup=get_add_product_cancel_keyboard()
        )

    await callback.answer()


@router.callback_query(F.data == "preorder_admin:new_category", AddPreorderProduct.category)
async def request_new_category(callback: CallbackQuery):
    """Запрос ввода новой категории"""
    await callback.message.edit_text(
        "Введите новую категорию:",
        reply_markup=get_add_product_cancel_keyboard()
    )
    await callback.answer()


@router.message(AddPreorderProduct.category)
async def process_category(message: Message, state: FSMContext):
    """Обработка категории"""
    await state.update_data(category=message.text)
    await state.set_state(AddPreorderProduct.product_name)

    # Получаем существующие названия товаров
    product_names = preorder_db.get_all_product_names()

    if product_names:
        await message.answer(
            f"Категория: {message.text}\n\nВыберите существующее название или введите новое:",
            reply_markup=get_product_name_selection_keyboard(product_names)
        )
    else:
        await message.answer(
            f"Категория: {message.text}\n\nВведите название товара:",
            reply_markup=get_add_product_cancel_keyboard()
        )


@router.callback_query(F.data.startswith("preorder_admin:select_product:"), AddPreorderProduct.product_name)
async def select_existing_product(callback: CallbackQuery, state: FSMContext):
    """Выбор существующего названия товара"""
    product_name = callback.data.split(":", 3)[2]
    await state.update_data(product_name=product_name)
    await state.set_state(AddPreorderProduct.flavor)

    data = await state.get_data()
    await callback.message.edit_text(
        f"Категория: {data['category']}\n"
        f"Товар: {product_name}\n\n"
        "Введите вкус товара:",
        reply_markup=get_add_product_cancel_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:new_product", AddPreorderProduct.product_name)
async def request_new_product(callback: CallbackQuery, state: FSMContext):
    """Запрос ввода нового названия товара"""
    data = await state.get_data()
    await callback.message.edit_text(
        f"Категория: {data['category']}\n\nВведите новое название товара:",
        reply_markup=get_add_product_cancel_keyboard()
    )
    await callback.answer()


@router.message(AddPreorderProduct.product_name)
async def process_product_name(message: Message, state: FSMContext):
    """Обработка названия товара"""
    await state.update_data(product_name=message.text)
    await state.set_state(AddPreorderProduct.flavor)

    data = await state.get_data()
    await message.answer(
        f"Категория: {data['category']}\n"
        f"Товар: {message.text}\n\n"
        "Введите вкус товара:",
        reply_markup=get_add_product_cancel_keyboard()
    )


@router.message(AddPreorderProduct.flavor)
async def process_flavor(message: Message, state: FSMContext):
    """Обработка вкуса"""
    await state.update_data(flavor=message.text)
    await state.set_state(AddPreorderProduct.description)

    data = await state.get_data()
    await message.answer(
        f"Категория: {data['category']}\n"
        f"Товар: {data['product_name']}\n"
        f"Вкус: {message.text}\n\n"
        "Введите описание товара или пропустите этот шаг:",
        reply_markup=get_skip_step_keyboard()
    )


@router.callback_query(F.data == "preorder_admin:skip_step", AddPreorderProduct.description)
async def skip_description(callback: CallbackQuery, state: FSMContext):
    """Пропустить описание"""
    await state.update_data(description=None)
    await state.set_state(AddPreorderProduct.price)

    data = await state.get_data()
    await callback.message.edit_text(
        f"Категория: {data['category']}\n"
        f"Товар: {data['product_name']}\n"
        f"Вкус: {data['flavor']}\n\n"
        "Введите цену товара или пропустите этот шаг:",
        reply_markup=get_skip_step_keyboard()
    )
    await callback.answer()


@router.message(AddPreorderProduct.description)
async def process_description(message: Message, state: FSMContext):
    """Обработка описания"""
    await state.update_data(description=message.text)
    await state.set_state(AddPreorderProduct.price)

    data = await state.get_data()
    await message.answer(
        f"Категория: {data['category']}\n"
        f"Товар: {data['product_name']}\n"
        f"Вкус: {data['flavor']}\n\n"
        "Введите цену товара или пропустите этот шаг:",
        reply_markup=get_skip_step_keyboard()
    )


@router.callback_query(F.data == "preorder_admin:skip_step", AddPreorderProduct.price)
async def skip_price(callback: CallbackQuery, state: FSMContext):
    """Пропустить цену"""
    await state.update_data(price=None)
    await state.set_state(AddPreorderProduct.expected_date)

    data = await state.get_data()
    await callback.message.edit_text(
        f"Категория: {data['category']}\n"
        f"Товар: {data['product_name']}\n"
        f"Вкус: {data['flavor']}\n\n"
        "Введите ожидаемую дату поставки (формат: ДД.ММ.ГГГГ) или пропустите:",
        reply_markup=get_skip_step_keyboard()
    )
    await callback.answer()


@router.message(AddPreorderProduct.price)
async def process_price(message: Message, state: FSMContext):
    """Обработка цены"""
    try:
        price = float(message.text)
        await state.update_data(price=price)
        await state.set_state(AddPreorderProduct.expected_date)

        data = await state.get_data()
        await message.answer(
            f"Категория: {data['category']}\n"
            f"Товар: {data['product_name']}\n"
            f"Вкус: {data['flavor']}\n\n"
            "Введите ожидаемую дату поставки (формат: ДД.ММ.ГГГГ) или пропустите:",
            reply_markup=get_skip_step_keyboard()
        )
    except ValueError:
        await message.answer(
            "❌ Неверный формат цены. Введите число:",
            reply_markup=get_skip_step_keyboard()
        )


@router.callback_query(F.data == "preorder_admin:skip_step", AddPreorderProduct.expected_date)
async def skip_date(callback: CallbackQuery, state: FSMContext):
    """Пропустить дату"""
    await state.update_data(expected_date=None)
    await state.set_state(AddPreorderProduct.image)

    await callback.message.edit_text(
        "Отправьте изображение товара или пропустите:",
        reply_markup=get_skip_step_keyboard()
    )
    await callback.answer()


@router.message(AddPreorderProduct.expected_date)
async def process_expected_date(message: Message, state: FSMContext):
    """Обработка даты поставки"""
    try:
        # Парсим дату
        date = datetime.strptime(message.text, '%d.%m.%Y')
        await state.update_data(expected_date=date.strftime('%Y-%m-%d'))
        await state.set_state(AddPreorderProduct.image)
        await message.answer(
            "Отправьте изображение товара или пропустите:",
            reply_markup=get_skip_step_keyboard()
        )
    except ValueError:
        await message.answer(
            "❌ Неверный формат даты. Используйте формат ДД.ММ.ГГГГ:",
            reply_markup=get_skip_step_keyboard()
        )


@router.message(AddPreorderProduct.image, F.photo)
async def process_image(message: Message, state: FSMContext):
    """Обработка изображения"""
    # Создаем папку для изображений, если её нет
    os.makedirs("product_images", exist_ok=True)

    # Сохраняем изображение
    photo = message.photo[-1]
    file_info = await message.bot.get_file(photo.file_id)
    file_path = f"product_images/preorder_{photo.file_id}.jpg"

    await message.bot.download_file(file_info.file_path, file_path)

    await state.update_data(image_path=file_path)

    # Отправляем новое сообщение вместо редактирования
    new_message = await message.answer("Подготовка данных...")
    await show_product_summary(new_message, state)


@router.callback_query(F.data == "preorder_admin:skip_step", AddPreorderProduct.image)
async def skip_image(callback: CallbackQuery, state: FSMContext):
    """Пропустить добавление изображения"""
    await state.update_data(image_path=None)
    # Создаем новое сообщение для показа сводки
    await callback.message.delete()
    new_message = await callback.message.answer("Подготовка данных...")
    await show_product_summary(new_message, state)
    await callback.answer()


async def show_product_summary(message: Message, state: FSMContext):
    """Показать сводку добавляемого товара"""
    data = await state.get_data()

    summary = "<b>Проверьте данные товара:</b>\n\n"
    summary += f"📦 Категория: {data['category']}\n"
    summary += f"🛍️ Товар: {data['product_name']}\n"
    summary += f"🍓 Вкус: {data['flavor']}\n"

    if data.get('description'):
        summary += f"📝 Описание: {data['description']}\n"
    else:
        summary += "📝 Описание: Не указано\n"

    if data.get('price'):
        summary += f"💰 Цена: {data['price']} ₽\n"
    else:
        summary += "💰 Цена: Не указана\n"

    if data.get('expected_date'):
        expected_date = datetime.strptime(data['expected_date'], '%Y-%m-%d').strftime('%d.%m.%Y')
        summary += f"📅 Дата поставки: {expected_date}\n"
    else:
        summary += "📅 Дата поставки: Не указана\n"

    summary += f"🖼️ Изображение: {'Добавлено' if data.get('image_path') else 'Не добавлено'}"

    await state.set_state(AddPreorderProduct.confirm)
    await message.edit_text(
        summary,
        reply_markup=get_confirm_add_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(F.data == "preorder_admin:edit_summary", AddPreorderProduct.confirm)
async def show_edit_menu(callback: CallbackQuery):
    """Показать меню редактирования"""
    await callback.message.edit_text(
        "Выберите поле для редактирования:",
        reply_markup=get_edit_field_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("preorder_admin:edit:"), AddPreorderProduct.confirm)
async def edit_field(callback: CallbackQuery, state: FSMContext):
    """Редактировать конкретное поле"""
    field = callback.data.split(":", 2)[2]
    await state.set_state(AddPreorderProduct.edit_field)
    await state.update_data(editing_field=field)

    field_names = {
        'category': 'категорию',
        'product_name': 'название товара',
        'flavor': 'вкус',
        'description': 'описание',
        'price': 'цену',
        'expected_date': 'дату поставки (ДД.ММ.ГГГГ)',
        'image': 'изображение'
    }

    if field == 'image':
        await callback.message.edit_text(
            f"Отправьте новое изображение или пропустите:",
            reply_markup=get_skip_step_keyboard()
        )
    else:
        text = f"Введите новое значение для поля '{field_names.get(field, field)}':"
        if field in ['description', 'price', 'expected_date']:
            text += "\n\nИли пропустите этот шаг:"
            keyboard = get_skip_step_keyboard()
        else:
            keyboard = get_add_product_cancel_keyboard()

        await callback.message.edit_text(text, reply_markup=keyboard)

    await callback.answer()


@router.message(AddPreorderProduct.edit_field)
async def process_edit_field(message: Message, state: FSMContext):
    """Обработка редактирования поля"""
    data = await state.get_data()
    field = data.get('editing_field')

    if field == 'price':
        try:
            value = float(message.text)
        except ValueError:
            await message.answer("❌ Неверный формат цены. Введите число:")
            return
    elif field == 'expected_date':
        try:
            date = datetime.strptime(message.text, '%d.%m.%Y')
            value = date.strftime('%Y-%m-%d')
        except ValueError:
            await message.answer("❌ Неверный формат даты. Используйте формат ДД.ММ.ГГГГ:")
            return
    else:
        value = message.text

    await state.update_data({field: value})
    await state.set_state(AddPreorderProduct.confirm)

    # Отправляем новое сообщение вместо редактирования
    new_message = await message.answer("Обновление данных...")
    await show_product_summary(new_message, state)


@router.message(AddPreorderProduct.edit_field, F.photo)
async def process_edit_image(message: Message, state: FSMContext):
    """Обработка редактирования изображения"""
    # Создаем папку для изображений, если её нет
    os.makedirs("product_images", exist_ok=True)

    # Удаляем старое изображение, если оно было
    data = await state.get_data()
    if data.get('image_path') and os.path.exists(data['image_path']):
        os.remove(data['image_path'])

    # Сохраняем новое изображение
    photo = message.photo[-1]
    file_info = await message.bot.get_file(photo.file_id)
    file_path = f"product_images/preorder_{photo.file_id}.jpg"

    await message.bot.download_file(file_info.file_path, file_path)

    await state.update_data(image_path=file_path)
    await state.set_state(AddPreorderProduct.confirm)

    # Отправляем новое сообщение вместо редактирования
    new_message = await message.answer("Обновление данных...")
    await show_product_summary(new_message, state)


@router.callback_query(F.data == "preorder_admin:skip_step", AddPreorderProduct.edit_field)
async def skip_edit_field(callback: CallbackQuery, state: FSMContext):
    """Пропустить редактирование поля"""
    data = await state.get_data()
    field = data.get('editing_field')

    if field in ['description', 'price', 'expected_date', 'image']:
        await state.update_data({field: None})

    await state.set_state(AddPreorderProduct.confirm)
    await callback.message.delete()
    new_message = await callback.message.answer("Подготовка данных...")
    await show_product_summary(new_message, state)
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:back_to_confirm")
async def back_to_confirm(callback: CallbackQuery, state: FSMContext):
    """Вернуться к подтверждению"""
    await state.set_state(AddPreorderProduct.confirm)
    await show_product_summary(callback.message, state)
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:confirm_add", AddPreorderProduct.confirm)
async def confirm_add_product(callback: CallbackQuery, state: FSMContext):
    """Подтвердить добавление товара"""
    data = await state.get_data()

    product_id = preorder_db.add_preorder_product(
        category=data['category'],
        product_name=data['product_name'],
        flavor=data['flavor'],
        description=data.get('description'),
        price=data.get('price'),
        expected_date=data.get('expected_date'),
        image_path=data.get('image_path')
    )

    if product_id:
        await callback.message.edit_text(
            "✅ Товар успешно добавлен!",
            reply_markup=get_preorder_admin_menu_keyboard()
        )
    else:
        await callback.message.edit_text(
            "❌ Ошибка при добавлении товара",
            reply_markup=get_preorder_admin_menu_keyboard()
        )

    await state.clear()
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:cancel_add")
async def cancel_add_product(callback: CallbackQuery, state: FSMContext):
    """Отменить добавление товара"""
    # Удаляем сохраненное изображение, если оно было
    data = await state.get_data()
    if data.get('image_path') and os.path.exists(data['image_path']):
        os.remove(data['image_path'])

    await state.clear()
    await callback.message.edit_text(
        "❌ Добавление товара отменено",
        reply_markup=get_preorder_admin_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:list")
async def show_products_list(callback: CallbackQuery):
    """Показать список товаров для предзаказа"""
    await display_products_list(callback, page=1)


@router.callback_query(F.data.startswith("preorder_admin:list_page:"))
async def show_products_list_page(callback: CallbackQuery):
    """Показать конкретную страницу списка товаров"""
    page = int(callback.data.split(":", 2)[2])
    await display_products_list(callback, page=page)


async def display_products_list(callback: CallbackQuery, page: int):
    """Отобразить список товаров с пагинацией"""
    data = preorder_db.get_all_preorder_products(page=page)

    if not data['items']:
        await callback.message.edit_text(
            "📋 Список товаров пуст",
            reply_markup=get_preorder_admin_menu_keyboard()
        )
    else:
        text = "📋 <b>Товары для предзаказа:</b>\n\n"
        if data['total_pages'] > 1:
            text += f"Страница {page} из {data['total_pages']}\n\n"
        text += "Нажмите на товар для подробной информации:"

        await callback.message.edit_text(
            text,
            reply_markup=get_preorder_products_list_keyboard(
                data['items'], page, data['total_pages']
            ),
            parse_mode="HTML"
        )

    await callback.answer()


@router.callback_query(F.data.startswith("preorder_admin:view:"))
async def view_product_details(callback: CallbackQuery):
    """Показать детали товара"""
    product_id = int(callback.data.split(":", 2)[2])

    # Получаем информацию о товаре
    product = preorder_db.get_product_by_id(product_id)

    if not product:
        await callback.answer("Товар не найден", show_alert=True)
        return

    text = f"<b>{product['product_name']}</b>\n\n"
    text += f"📦 Категория: {product['category']}\n"
    text += f"🍓 Вкус: {product['flavor']}\n"

    if product['description']:
        text += f"📝 Описание: {product['description']}\n"

    if product['price']:
        text += f"💰 Цена: {product['price']} ₽\n"
    else:
        text += "💰 Цена: Не указана\n"

    if product['expected_date']:
        expected_date = datetime.strptime(product['expected_date'], '%Y-%m-%d').strftime('%d.%m.%Y')
        text += f"📅 Дата поставки: {expected_date}\n"
    else:
        text += "📅 Дата поставки: Не указана\n"

    text += f"👁 Просмотров: {product['views']}\n"
    text += f"📊 Предзаказов: {product.get('preorder_count', 0)}\n"
    text += f"🖼️ Изображение: {'Есть' if product['image_path'] else 'Нет'}"

    await callback.message.edit_text(
        text,
        reply_markup=get_product_admin_keyboard(product_id),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("preorder_admin:delete:"))
async def request_delete_confirmation(callback: CallbackQuery):
    """Запросить подтверждение удаления"""
    product_id = int(callback.data.split(":", 2)[2])

    await callback.message.edit_text(
        "⚠️ Вы уверены, что хотите удалить этот товар?\n"
        "Все предзаказы на него будут отменены.",
        reply_markup=get_confirm_delete_keyboard(product_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("preorder_admin:confirm_delete:"))
async def confirm_delete_product(callback: CallbackQuery):
    """Подтвердить удаление товара"""
    product_id = int(callback.data.split(":", 2)[2])

    # Получаем информацию о товаре перед удалением
    product = preorder_db.get_product_by_id(product_id)

    # Получаем список пользователей с предзаказом
    users_to_notify = preorder_db.get_users_with_preorder(product_id)

    if preorder_db.delete_preorder_product(product_id):
        # Отправляем уведомления пользователям
        if users_to_notify and product:
            bot: Bot = callback.bot
            for user_id in users_to_notify:
                try:
                    await bot.send_message(
                        user_id,
                        f"⚠️ <b>Важное уведомление</b>\n\n"
                        f"К сожалению, товар <b>{product['product_name']} ({product['flavor']})</b> "
                        f"был исключен из ближайшей поставки.\n\n"
                        f"Ваш предзаказ на этот товар был автоматически отменен.\n\n"
                        f"Рекомендуем посмотреть другие товары в разделе предзаказов!",
                        parse_mode="HTML",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(
                                text="🛍️ Посмотреть другие товары",
                                callback_data="profile:preorders"
                            )
                        ]])
                    )
                except Exception as e:
                    logger.error(f"Не удалось отправить уведомление пользователю {user_id}: {e}")

        await callback.message.edit_text(
            "✅ Товар успешно удален\n"
            f"📨 Уведомлений отправлено: {len(users_to_notify)}",
            reply_markup=get_preorder_admin_menu_keyboard()
        )
    else:
        await callback.message.edit_text(
            "❌ Ошибка при удалении товара",
            reply_markup=get_preorder_admin_menu_keyboard()
        )

    await callback.answer()


@router.callback_query(F.data == "preorder_admin:stats")
async def show_preorder_stats(callback: CallbackQuery):
    """Показать статистику предзаказов"""
    await display_stats(callback, page=1)


@router.callback_query(F.data.startswith("preorder_admin:stats_page:"))
async def show_stats_page(callback: CallbackQuery):
    """Показать конкретную страницу статистики"""
    page = int(callback.data.split(":", 2)[2])
    await display_stats(callback, page=page)


async def display_stats(callback: CallbackQuery, page: int):
    """Отобразить статистику с пагинацией"""
    # Получаем все товары для статистики
    all_products_data = preorder_db.get_all_preorder_products(page=1, per_page=1000)
    all_products = all_products_data['items']

    total_products = len(all_products)
    total_preorders = sum(p.get('preorder_count', 0) for p in all_products)
    total_views = sum(p['views'] for p in all_products)

    # Топ товаров по предзаказам
    products_with_orders = [p for p in all_products if p.get('preorder_count', 0) > 0]
    top_products = sorted(products_with_orders, key=lambda x: x.get('preorder_count', 0), reverse=True)

    # Пагинация для топа
    items_per_page = 10
    total_pages = max(1, (len(top_products) + items_per_page - 1) // items_per_page)
    start_idx = (page - 1) * items_per_page
    end_idx = start_idx + items_per_page
    page_products = top_products[start_idx:end_idx]

    text = (
        "📊 <b>Статистика предзаказов</b>\n\n"
        f"📦 Всего товаров: {total_products}\n"
        f"📋 Всего предзаказов: {total_preorders}\n"
        f"👁 Всего просмотров: {total_views}\n\n"
    )

    if page_products:
        text += "<b>🏆 Топ товаров по предзаказам:</b>\n"
        for i, product in enumerate(page_products, start_idx + 1):
            text += f"{i}. {product['product_name']} ({product['flavor']}) - {product['preorder_count']} заказов\n"

        if total_pages > 1:
            text += f"\nСтраница {page} из {total_pages}"
    elif page == 1:
        text += "\nПока нет товаров с предзаказами."

    cancellation_stats = preorder_db.get_cancellation_stats()

    if cancellation_stats['total'] > 0:
        text += f"\n\n📊 <b>Статистика отказов:</b>\n"
        text += f"Всего отказов: {cancellation_stats['total']}\n\n"

        text += "<b>По причинам:</b>\n"
        reason_names = {
            'financial': '💰 Финансовые обстоятельства',
            'too_long': '⏰ Слишком долго ждать',
            'found_elsewhere': '🔄 Нашёл в другом месте',
            'changed_mind': '❌ Передумал',
            'other': '📝 Другая причина'
        }

        for reason, count in cancellation_stats['by_reason']:
            reason_text = reason_names.get(reason, reason)
            percentage = (count / cancellation_stats['total']) * 100
            text += f"• {reason_text}: {count} ({percentage:.1f}%)\n"

    await callback.message.edit_text(
        text,
        reply_markup=get_stats_keyboard(page, total_pages),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:current_page")
@router.callback_query(F.data == "preorder_admin:current_stats_page")
async def handle_current_page_click(callback: CallbackQuery):
    """Обработка клика на текущую страницу"""
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:bulk_upload")
async def start_bulk_upload(callback: CallbackQuery, state: FSMContext):
    """Начать массовую загрузку товаров"""
    await state.set_state(BulkUploadProducts.upload_file)

    await callback.message.edit_text(
        "📤 <b>Массовая загрузка товаров</b>\n\n"
        "Для загрузки товаров используйте Excel файл со следующими колонками:\n"
        "• Категория\n"
        "• Название товара\n"
        "• Вкус\n"
        "• Описание (необязательно)\n"
        "• Цена (необязательно)\n"
        "• Дата поставки (формат: ДД.ММ.ГГГГ, необязательно)\n\n"
        "Отправьте Excel файл или скачайте шаблон:",
        reply_markup=get_bulk_upload_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:download_template")
async def download_template(callback: CallbackQuery):
    """Скачать шаблон Excel файла"""
    # Создаем шаблон Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары для предзаказа"

    # Заголовки
    headers = ["Категория", "Название товара", "Вкус", "Описание", "Цена", "Дата поставки"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Примеры данных
    examples = [
        ["Жидкости", "Brusko", "Клубника", "Премиальная жидкость", "350", "25.12.2024"],
        ["Устройства", "OXVA Xlim", "Черный", "Компактный под-мод", "1500", "30.12.2024"],
        ["Картриджи", "Voopoo PnP", "0.8 Ом", "Сменный испаритель", "", ""]
    ]

    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Устанавливаем ширину колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Отправляем файл
    await callback.message.answer_document(
        BufferedInputFile(
            buffer.read(),
            filename="template_preorder_products.xlsx"
        ),
        caption="📥 Шаблон для массовой загрузки товаров\n\n"
                "Заполните файл и отправьте его обратно."
    )
    await callback.answer()


@router.message(BulkUploadProducts.upload_file, F.document)
async def process_excel_file(message: Message, state: FSMContext):
    """Обработка Excel файла"""
    document = message.document

    # Проверяем расширение файла
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer(
            "❌ Неверный формат файла. Отправьте Excel файл (.xlsx или .xls)",
            reply_markup=get_bulk_upload_keyboard()
        )
        return

    # Скачиваем файл
    file_info = await message.bot.get_file(document.file_id)
    file_bytes = BytesIO()
    await message.bot.download_file(file_info.file_path, file_bytes)
    file_bytes.seek(0)

    try:
        # Читаем Excel файл
        wb = openpyxl.load_workbook(file_bytes, read_only=True)
        ws = wb.active

        products = []
        errors = []

        # Пропускаем заголовок
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row_idx, row in enumerate(rows, 2):
            if not any(row):  # Пропускаем пустые строки
                continue

            try:
                category = str(row[0]).strip() if row[0] else None
                product_name = str(row[1]).strip() if row[1] else None
                flavor = str(row[2]).strip() if row[2] else None
                description = str(row[3]).strip() if row[3] and str(row[3]).strip() else None

                # Обработка цены
                price = None
                if row[4]:
                    try:
                        price = float(str(row[4]).replace(',', '.'))
                    except ValueError:
                        errors.append(f"Строка {row_idx}: неверный формат цены")
                        continue

                # Обработка даты
                expected_date = None
                if row[5]:
                    if isinstance(row[5], datetime):
                        expected_date = row[5].strftime('%Y-%m-%d')
                    else:
                        try:
                            date_str = str(row[5]).strip()
                            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                            expected_date = date_obj.strftime('%Y-%m-%d')
                        except ValueError:
                            errors.append(f"Строка {row_idx}: неверный формат даты")
                            continue

                # Проверка обязательных полей
                if not all([category, product_name, flavor]):
                    errors.append(f"Строка {row_idx}: не заполнены обязательные поля")
                    continue

                products.append({
                    'category': category,
                    'product_name': product_name,
                    'flavor': flavor,
                    'description': description,
                    'price': price,
                    'expected_date': expected_date
                })

            except Exception as e:
                errors.append(f"Строка {row_idx}: ошибка обработки - {str(e)}")

        wb.close()

        if not products:
            await message.answer(
                "❌ Не удалось загрузить ни одного товара.\n" +
                ("\nОшибки:\n" + "\n".join(errors) if errors else ""),
                reply_markup=get_preorder_admin_menu_keyboard()
            )
            await state.clear()
            return

        # Сохраняем данные в состоянии
        await state.update_data(products=products, errors=errors)
        await state.set_state(BulkUploadProducts.confirm)

        # Показываем сводку
        summary = f"📊 <b>Результат обработки файла:</b>\n\n"
        summary += f"✅ Товаров готово к загрузке: {len(products)}\n"
        if errors:
            summary += f"❌ Ошибок: {len(errors)}\n\n"
            summary += "<b>Список ошибок:</b>\n"
            for error in errors[:10]:  # Показываем первые 10 ошибок
                summary += f"• {error}\n"
            if len(errors) > 10:
                summary += f"\n... и еще {len(errors) - 10} ошибок"

        await message.answer(
            summary,
            reply_markup=get_bulk_upload_confirm_keyboard(),
            parse_mode="HTML"
        )

    except Exception as e:
        logger.error(f"Ошибка при обработке Excel файла: {e}")
        await message.answer(
            f"❌ Ошибка при обработке файла: {str(e)}",
            reply_markup=get_preorder_admin_menu_keyboard()
        )
        await state.clear()


@router.callback_query(F.data == "preorder_admin:confirm_bulk", BulkUploadProducts.confirm)
async def confirm_bulk_upload(callback: CallbackQuery, state: FSMContext):
    """Подтвердить массовую загрузку"""
    data = await state.get_data()
    products = data.get('products', [])

    success_count = 0
    failed_count = 0

    for product in products:
        try:
            product_id = preorder_db.add_preorder_product(
                category=product['category'],
                product_name=product['product_name'],
                flavor=product['flavor'],
                description=product['description'],
                price=product['price'],
                expected_date=product['expected_date']
            )
            if product_id:
                success_count += 1
            else:
                failed_count += 1
        except Exception as e:
            logger.error(f"Ошибка при добавлении товара: {e}")
            failed_count += 1

    result_text = f"✅ <b>Массовая загрузка завершена!</b>\n\n"
    result_text += f"Успешно загружено: {success_count} товаров\n"
    if failed_count > 0:
        result_text += f"Не удалось загрузить: {failed_count} товаров"

    await callback.message.edit_text(
        result_text,
        reply_markup=get_preorder_admin_menu_keyboard(),
        parse_mode="HTML"
    )

    await state.clear()
    await callback.answer()


@router.callback_query(F.data == "preorder_admin:cancel_bulk")
async def cancel_bulk_upload(callback: CallbackQuery, state: FSMContext):
    """Отменить массовую загрузку"""
    await state.clear()
    await callback.message.edit_text(
        "❌ Массовая загрузка отменена",
        reply_markup=get_preorder_admin_menu_keyboard()
    )
    await callback.answer()
